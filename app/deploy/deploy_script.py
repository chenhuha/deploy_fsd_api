import os
import json
import yaml
import subprocess
import time
import shutil
from openpyxl import load_workbook
from common import types, utils, constants
from deploy.preview import Preview
from flask import current_app
from uuid import uuid1
from threading import Thread
from deploy.node_base import Node
from models.deploy_history import DeployHistoryModel
from models.load_info import LoadInfoModel
from models.upgrade_history import UpgradeHistoryModel
from models.deploy_status import DeployStatusModel
from models.extend_history import ExtendHistoryModel


class DeployScript(Preview, Node):
    def __init__(self):
        super().__init__()
        self.global_vars_path = os.path.join(
            current_app.config['ETC_EXAMPLE_PATH'], 'global_vars.yaml')
        self.ceph_globals_path = os.path.join(
            current_app.config['ETC_EXAMPLE_PATH'], 'ceph-globals.yaml')
        self.hosts_path = os.path.join(
            current_app.config['ETC_EXAMPLE_PATH'], 'hosts')
        self.db_path = current_app.config['DB_NAME']
        self.deploy_info_path = os.path.join(
            current_app.config['DEPLOY_HOME'], 'deploy_node_info.xlsx')
        self.bak_path = os.path.join(current_app.config['ETC_EXAMPLE_PATH'], time.strftime(
            '%Y-%m-%d', time.localtime(time.time())) + '_example_bak/')

        self.deploy_history_model = DeployHistoryModel()
        self.deploy_status_model = DeployStatusModel()

    def post(self):
        preview_info = self.get_preview_from_request()
        config_file = self.file_conversion(preview_info)
        for config in config_file:
            file_path = os.path.join(
                current_app.config['ETC_EXAMPLE_PATH'], config['shellName'])
            self._config_bak(
                current_app.config['ETC_EXAMPLE_PATH'], config['shellName'])
            with open(file_path, 'w', encoding='UTF-8') as f:
                f.write(config['shellContent'])
        self.control_deploy(preview_info)
        return types.DataModel().model(code=0, data="")

    def control_deploy(self, previews):
        if self.deploy_history_model.get_deploy_history():
            deploy_type = "retry"
        else:
            deploy_type = "first"
        ceph_flag = previews['common']['commonFixed']['cephServiceFlag']
        deploy_key = previews['key']
        deploy_uuid = str(uuid1())
        results = types.DataModel().history_deploy_model(
            paramsJson=json.dumps(previews),
            uuid=deploy_uuid,
            startTime=int(time.time() * 1000),
            endtime=int(time.time() * 1000)
        )
        self._write_history_file(results)
        self._create_status_table()
        cmd = ['sh', current_app.config['SCRIPT_PATH'] + '/setup.sh',
               deploy_key, deploy_type, str(ceph_flag), str(deploy_uuid)]
        self._logger.info('deploy command: %s', cmd)
        results = subprocess.Popen(cmd, stdout=subprocess.PIPE)
        thread = Thread(target=self._shell_return_listen, args=(
            current_app._get_current_object(), results, previews, deploy_uuid, int(time.time() * 1000)))
        thread.start()

    def _config_bak(self, file_path, file_name):
        if not os.path.exists(self.bak_path):
            os.makedirs(self.bak_path)
        filepath = os.path.join(file_path, file_name)
        bak_filename = file_name + \
            time.strftime('%H-%M-%S', time.localtime(time.time())) + '_bak'
        bak_file_path = os.path.join(self.bak_path, bak_filename)
        shutil.copy(filepath,  bak_file_path)

    def _shell_return_listen(self, app, subprocess_1, previews, deploy_uuid, start_time):
        with app.app_context():
            subprocess_1.wait()
            status = self.deploy_status_model.get_deploy_last_status()
            if status:
                deploy_message = status[0]
                deploy_result = status[1]
            else:
                deploy_message = 'deploy faild.'
                deploy_result = 'false'
            results = types.DataModel().history_deploy_model(
                log=str(subprocess_1.stdout.read(), encoding='utf-8'),
                paramsJson=json.dumps(previews),
                uuid=deploy_uuid,
                startTime=start_time,
                endtime=int(time.time() * 1000),
                message=deploy_message,
                result=deploy_result
            )
            self._write_history_file(results)
            if deploy_result.lower() == 'true':
                self._write_node_info_csv(previews['nodes'])
                self._write_upgrade_file()
                self._write_extend_file(results)
                self.scp_deploy(previews['nodes'])

    def _net_info(self, cards):
        cards_list = []
        for card in cards:
            purpose = []
            if 'EXTRANET' in card['purpose']:
                purpose.append("业务网")
            if 'MANAGEMENT' in card['purpose']:
                purpose.append("管理网")
            if 'STORAGECLUSTER' in card['purpose']:
                purpose.append("存储集群网")
            if 'STORAGEPUBLIC' in card['purpose']:
                purpose.append("存储公网")
            cards_list.append([card['name'], card.get('ip', 'null'), ','.join(
                purpose), card['bond'], card['speed'], card.get('mode', 'null'), card.get('mtu', 'null'), card.get('pciid', 'null'), card.get('slaves', 'null')
            ])
        return cards_list

    def _storages_info(self, storages):
        hdd_storages_info = []
        ssd_storages_info = []
        storage_load_dict = self._load_storage()
        for storage in storages:
            if storage_load_dict:
                _bool, storgae_info = self._ssd_bool(
                    storage['name'], storage_load_dict)
            else:
                _bool = False
                storgae_info = {}
            if _bool:
                ssd_storages_info.append([
                    storage['name'],
                    self._storage_purpose_convert(storage['purpose']),
                    storage['size'],
                    storgae_info.get('model', 'null'),
                    storgae_info.get('partition', '[]'),
                    str(storage['cache2data'])
                ])
            else:
                hdd_storages_info.append([
                    storage['name'],
                    self._storage_purpose_convert(storage['purpose']),
                    storage['size'],
                    storgae_info.get('model', 'null'),
                    storgae_info.get('partition', '[]')
                ])
        return hdd_storages_info, ssd_storages_info

    def _storage_purpose_convert(self, purpose):
        if purpose == "SYSTEM":
            return "系统盘"
        elif purpose == "DATA":
            return "ceph数据盘"
        elif purpose == "CACHE":
            return "ceph缓存盘"
        elif purpose == "VOIDATA":
            return "VOI数据盘"

    def _ssd_bool(self, name, storage_load_dict):
        for node in storage_load_dict:
            for ssd in node['ssds']:
                if name == ssd.get('name'):
                    return True, ssd
            for hdd in node['hdds']:
                if name == hdd.get('name'):
                    return False, hdd
        return False, {}

    def _load_storage(self):
        try:
            model = LoadInfoModel()
            info = model.get_load_info()
            return json.loads(info)
        except Exception as e:
            return []

    def scp_deploy(self, nodes):
        cmds = []
        for node in nodes:
            for file in [self.db_path, self.global_vars_path, self.ceph_globals_path, self.hosts_path, self.deploy_info_path]:
                cmd = constants.COMMAND_SCP_FILE % (
                    current_app.config['NODE_PASS'], file, current_app.config['NODE_USER'], node['nodeIP'], file)
                cmds.append(cmd)
        try:
            for cmd in cmds:
                _, result, _ = utils.execute(cmd)
                self._logger.info(f"Execute command '{cmd}', result:{result}")
        except Exception as e:
            self._logger.error(
                f"Execute command to copy history is faild ,Because: {e}")

    def version(self):
        if os.path.exists('/etc/klcloud-release'):
            with open('/etc/klcloud-release', 'r') as f:
                version = f.read()
        else:
            with open(os.path.join(current_app.config['ETC_EXAMPLE_PATH'], 'global_vars.yaml'), 'r') as f:
                global_var = yaml.load(f.read(), Loader=yaml.FullLoader)
            if global_var['deploy_edu']:
                version = f"EDU-v{global_var['fsd_default_tag']}"
            else:
                version = f"COMM-v{global_var['fsd_default_tag']}"
            with open('/etc/klcloud-release', 'w') as f:
                f.write(version)

        return version.strip()

    def _write_node_info_csv(self, nodes):
        book = load_workbook(self.deploy_info_path)
        template_sheet = book['mould']
        for node in nodes:
            target_sheet = book.copy_worksheet(template_sheet)
            target_sheet.title = node['nodeName']
            target_sheet.cell(row=3, column=1, value=node['nodeName'])
            target_sheet.cell(row=3, column=2, value=node['nodeIP'])
            target_sheet.cell(
                row=3, column=3, value=','.join(node['nodeType']))
            target_sheet.cell(row=3, column=4, value='0.0.0.0')
            net_info = self._net_info(node['networkCards'])
            self._write_info_csv(net_info, target_sheet, 3, 6)
            hdd_info, ssd_info = self._storages_info(node['storages'])
            self._write_info_csv(hdd_info, target_sheet, 3, 16)
            self._write_info_csv(ssd_info, target_sheet, 3, 22)

        book.save(self.deploy_info_path)
        book.close()

    def _write_info_csv(self, infos, sheet, start_row, start_col):
        for row, info in enumerate(infos, start=start_row):
            for col, value in enumerate(info, start=start_col):
                sheet.cell(row=row, column=col, value=value)

    def _write_history_file(self, result):
        self.deploy_history_model.create_deploy_history_table()
        self.deploy_history_model.add_deploy_history(
            result['paramsJson'], result['log'], result['message'],
            result['uuid'], result['result'], result['startTime'],
            result['endtime'], result['key'])

    def _write_upgrade_file(self):
        version = self.version()
        model = UpgradeHistoryModel()
        model.create_upgrade_history_table()
        model.add_upgrade_history(
            '', version, "true", '', int(time.time() * 1000))

    def _write_extend_file(self, result):
        model = ExtendHistoryModel()
        model.create_extend_history_table()
        model.add_extend_history(
            result['paramsJson'], result['log'], 
            result['message'], result['result'], 
            result['startTime'], result['endtime']
        )
        
    def _create_status_table(self):
        self.deploy_status_model.create_deploy_status_table()
