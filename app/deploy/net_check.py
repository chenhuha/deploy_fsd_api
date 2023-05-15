import json
import os
import shutil
import paramiko
import openpyxl
from models.load_info import LoadInfoModel

from common import types
from deploy.node_base import Node
from flask import current_app
from flask_restful import reqparse, Resource


class NetCheck(Resource, Node):
    def __init__(self):
        super().__init__()
        self.nodes = self.get_nodes_from_request()
        self.node_list = self.get_info_with_from(self.nodes)

    def post(self):
        data = {}
        if len(self.nodes) == 1:
            data = self.single_node_data()
        else:
            data = self.multiple_nodes_data()

        # 保存数据到本地
        node_info_file = os.path.join(
            self.deploy_home, "deploy_node_info.xlsx")
        if os.path.isfile(node_info_file):
            os.remove(node_info_file)
        node_info_file = os.path.join(
            self.deploy_home, "deploy_node_info.xlsx")
        source_file = os.path.join(self.template_path, "deployExcel.xlsx")
        try:
            shutil.copyfile(source_file, node_info_file)
            self.write_data_to_excel(node_info_file, data)
        except Exception as e:
            self._logger.error('write data to excel error, %s', e)
        
        return types.DataModel().model(code=0, data=data)

    # 单节点
    def single_node_data(self):
        node = self.node_list[0]
        api_result = []
        storage_cluster_result = []
        storage_public_result = []

        if 'management' in node:
            status = 2 if int(node['management']['speed']) < 1000 else 0
            result = {
                'sourceIp': node['management']['ip'],
                'sourceHostname': node['hostname'],
                'destIp': node['management']['ip'],
                'destHostname': node['hostname'],
                'speed': node['management']['speed'],
                'realSpeed': '-',
                'mtu': node['management']['mtu'],
                'plr': '-',
                'status': status
            }
            api_result.append(result)

        if 'storage_cluster' in node:
            status = 2 if int(node['storage_cluster']['speed']) < 10000 else 0
            result = {
                'sourceIp': node['storage_cluster']['ip'],
                'sourceHostname': node['hostname'],
                'destIp': node['storage_cluster']['ip'],
                'destHostname': node['hostname'],
                'speed': node['storage_cluster']['speed'],
                'realSpeed': '-',
                'mtu': node['storage_cluster']['mtu'],
                'plr': '-',
                'status': status
            }
            storage_cluster_result.append(result)

        if 'storage_public' in node:
            status = 2 if int(node['storage_public']['speed']) < 10000 else 0
            result = {
                'sourceIp': node['storage_public']['ip'],
                'sourceHostname': node['hostname'],
                'destIp': node['storage_public']['ip'],
                'destHostname': node['hostname'],
                'speed': node['storage_public']['speed'],
                'realSpeed': '-',
                'mtu': node['storage_public']['mtu'],
                'plr': '-',
                'status': status
            }
            storage_public_result.append(result)

        return self.combine_results(api_result, storage_cluster_result, storage_public_result)

    # 多节点
    def multiple_nodes_data(self):
        api_result = []
        ceph_cluster_result = []
        ceph_public_result = []

        def output_format(current_node, node, node_ip, port, purpose):
            try:
                output = self.iperf3_client(
                    current_node['management']['ip'], node_ip, port)
                client_result, server_result = self.output_format_different_node(
                    output, purpose)
            except Exception as e:
                self._logger.error('get iperf3_client output failed, %s', e)
                client_result, server_result = self.output_format_null_node(current_node, node, purpose)
            return client_result, server_result

        # execute iperf3 client and collect results
        for i, current_node in enumerate(self.node_list):
            for j, node in enumerate(self.node_list[i:], i):
                if i == j:
                    api_result.append(
                        self.output_format_same_node(node, 'management'))
                    ceph_cluster_result.append(
                        self.output_format_same_node(node, 'storage_cluster'))
                    ceph_public_result.append(
                        self.output_format_same_node(node, 'storage_public'))
                else:
                    # prepare iperf3 server
                    for port in [5201, 5202, 5203]:
                        self.iperf3_server(node['management']['ip'], port)

                    api_result.extend(output_format(
                        current_node, node, node['management']['ip'], 5201, 'management'))
                    ceph_cluster_result.extend(output_format(
                        current_node, node, node['storage_cluster']['ip'], 5202, 'storage_cluster'))
                    ceph_public_result.extend(output_format(
                        current_node, node, node['storage_public']['ip'], 5203, 'storage_public'))

        return self.combine_results(api_result, ceph_cluster_result, ceph_public_result)

    # 处理前端传来的数据
    def get_info_with_from(self, nodes):
        node_list = []
        for node in nodes:
            node_info = {
                'hostname': node['nodeName'], 'nodeIP': node['nodeIP']}
            for card in node['cards']:
                purpose = card.get('purpose')
                if 'MANAGEMENT' in purpose:
                    node_info['management'] = {
                        'ip': card['ip'], 'speed': card['speed'], 'mtu': card['mtu']}
                if 'STORAGECLUSTER' in purpose:
                    node_info['storage_cluster'] = {
                        'ip': card['ip'], 'speed': card['speed'], 'mtu': card['mtu']}
                if 'STORAGEPUBLIC' in purpose:
                    node_info['storage_public'] = {
                        'ip': card['ip'], 'speed': card['speed'], 'mtu': card['mtu']}
            node_list.append(node_info)

        return node_list

    def iperf3_server(self, host_ip, port):
        cmd = f'iperf3 -s -J -1 -p {port}'
        with paramiko.SSHClient() as ssh:
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(host_ip, username=self.username,
                        password=self.password)
            ssh.exec_command(cmd)

    def iperf3_client(self, host_ip, server, port):
        cmd = f'iperf3 -c {server} -p {port} -t 3 -i 1 --get-server-output -J'
        with paramiko.SSHClient() as ssh:
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(host_ip, username=self.username,
                        password=self.password)
            _, stdout, _ = ssh.exec_command(cmd)
            data = stdout.read().decode()
        return data

    # 处理相同节点的数据
    def output_format_same_node(self, node, purpose):
        source_ip = node[purpose]['ip']
        source_hostname = node["hostname"]
        dest_ip = source_ip
        dest_hostname = source_hostname
        speed = node[purpose]['speed']
        real_speed = int(speed) / 8
        mtu = node[purpose]['mtu']
        plr = '0%'
        status = self._get_status(speed, real_speed, plr, source_ip, purpose)

        result = {
            'sourceIp': source_ip,
            'sourceHostname': source_hostname,
            'destIp': dest_ip,
            'destHostname': dest_hostname,
            'speed': speed,
            'realSpeed': real_speed,
            'mtu': mtu,
            'packetLossRate': plr,
            'status': status
        }

        return result

    # 处理 iperf 返回的数据
    def output_format_different_node(self, result, purpose):
        json_data = json.loads(result)

        local_host = json_data['start']['connected'][0]['local_host']
        remote_host = json_data['start']['connected'][0]['remote_host']
        client_bits_per_second = json_data['end']['sum_received']['bits_per_second']
        server_bits_per_second = json_data['server_output_json']['end']['sum_received']['bits_per_second']
        client_hostname, client_speed, client_mtu = self._get_node_property(
            local_host)
        server_hostname, server_speed, server_mtu = self._get_node_property(
            remote_host)
        client_real_speed = self._get_realSpeed(client_bits_per_second)
        server_real_speed = self._get_realSpeed(server_bits_per_second)
        plr = self._get_packet_loss_rate(local_host, remote_host)
        client_status = self._get_status(
            client_speed, client_real_speed, plr, local_host, purpose)
        server_status = self._get_status(
            server_speed, server_real_speed, plr, remote_host, purpose)

        client_result = {
            'sourceIp': local_host,
            'sourceHostname': client_hostname,
            'destIp': remote_host,
            'destHostname': server_hostname,
            'speed': client_speed,
            'realSpeed': client_real_speed,
            'mtu': client_mtu,
            'plr': plr,
            'status': client_status
        }
        server_result = {
            'sourceIp': remote_host,
            'sourceHostname': server_hostname,
            'destIp': local_host,
            'destHostname': client_hostname,
            'speed': server_speed,
            'realSpeed': server_real_speed,
            'mtu': server_mtu,
            'plr': plr,
            'status': server_status
        }

        return client_result, server_result

    def output_format_null_node(self, current_node, node, purpose):
        source_ip = current_node[purpose]['ip']
        source_hostname = current_node["hostname"]
        dest_ip = node[purpose]['ip']
        dest_hostname = node["hostname"]
        speed = node[purpose]['speed']
        real_speed = 0
        mtu = node[purpose]['mtu']
        packet_loss_rate = '-'
        status = 1

        client_result = {
            'sourceIp': source_ip,
            'sourceHostname': source_hostname,
            'destIp': dest_ip,
            'destHostname': dest_hostname,
            'speed': speed,
            'realSpeed': real_speed,
            'mtu': mtu,
            'plr': packet_loss_rate,
            'status': status
        }

        server_result = {
            'sourceIp': dest_ip,
            'sourceHostname': dest_hostname,
            'destIp': source_ip,
            'destHostname': source_hostname,
            'speed': speed,
            'realSpeed': real_speed,
            'mtu': mtu,
            'plr': packet_loss_rate,
            'status': status
        }
        
        return client_result, server_result

    def combine_results(self, api_result, ceph_cluster_result, ceph_public_result):
        return {
            'apiResult': api_result,
            'cephClusterResult': ceph_cluster_result,
            'cephPublicResult': ceph_public_result
        }

    def _get_packet_loss_rate(self, local_host, remote_host):
        cmd = f'ping -c 20 -i 0.1 -W 1 -q {remote_host}'
        with paramiko.SSHClient() as ssh:
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(local_host, username=self.username,
                        password=self.password)
            _, stdout, _ = ssh.exec_command(cmd)
            output = stdout.read().decode()
            packet_loss = output.split(',')[-2].strip().split()[0]

        return packet_loss

    def _get_node_property(self, node_ip):
        for node in self.node_list:
            for key in ['management', 'storage_cluster', 'storage_public']:
                if node[key]['ip'] == node_ip:
                    return node['hostname'], node[key]['speed'], node[key]['mtu']

        return None, 0, 0

    def _get_realSpeed(self, bits_per_second):
        return round(bits_per_second / 1000000 / 8, 2)

    def _get_status(self, speed, real_speed, plr, node_ip, purpose):
        if plr != '0%':
            self._logger.info('plr not is 0%')
            return 1

        if purpose == 'management' and int(speed) < 1000:
            self._logger.warn(
                f'The speed of the management card < 1000, {node_ip}')
            return 2

        if purpose == 'storage_cluster' and int(speed) < 10000:
            self._logger.warn(
                f'The speed of the storage_cluster card < 10000, {node_ip}')
            return 2

        if purpose == 'storage_public' and int(speed) < 10000:
            self._logger.warn(
                f'The speed of the storage_cluster card < 10000, {node_ip}')
            return 2

        if int(real_speed) < int(speed) / 8 * 0.5:
            self._logger.warn(
                f'The real-time bandwidth is less than 50% of the standard, {node_ip}')
            return 2

        return 0

    # 数据持久化
    def write_data_to_excel(self, filepath, data):
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active
        start_num = 3

        status_map = {0: '正常', 1: '异常', 2: '警告'}
        card_name_map = {'apiResult': '管理网',
                         'cephPublicResult': '存储公网',
                         'cephClusterResult': '存储集群网'}
        for card_name, card_data in data.items():
            card_title = card_name_map.get(card_name, card_name)
            for item in card_data:
                worksheet.cell(row=start_num, column=1, value=card_title)
                worksheet.cell(row=start_num, column=2, value=item['sourceIp'])
                worksheet.cell(row=start_num, column=3, value=item['destIp'])
                worksheet.cell(row=start_num, column=4,
                               value=status_map.get(item['status'], ''))
                worksheet.cell(row=start_num, column=5, value=str(item))
                start_num += 1
        workbook.save(filepath)


class NetCheckCommon(NetCheck):
    def __init__(self):
        nodes = self.get_nodes_from_request()
        cards = self.get_cards_from_request()
        self.nodes = self.uniform_format_with_nodes(nodes, cards)
        super().__init__()

    def get_cards_from_request(self):
        parser = reqparse.RequestParser()
        parser.add_argument('cards', required=True, location='json',
                            type=list, help='The cards field does not exist')

        return parser.parse_args()['cards']

    def uniform_format_with_nodes(self, nodes, cards):
        load_info = self.load_storage()
        for node in nodes:
            node_cards = []
            node_ip = node['nodeIP']
            for card in cards:
                if 'MANAGEMENT' in card['purpose']:
                    card['ip'] = node_ip
                elif 'STORAGEPUBLIC' in card['purpose'] or 'STORAGECLUSTER' in card['purpose']:
                    card_ip = self.get_card_ip(
                        load_info, node_ip, card['name'])
                    card['ip'] = card_ip if card_ip else node_ip
                node_cards.append(card.copy())
            node['cards'] = node_cards
        return nodes

    def get_card_ip(self, load_info, node_ip, card_name):
        for node in load_info:
            if node['nodeIP'] == node_ip:
                for card in node['cards']:
                    if card['name'] == card_name:
                        return card['ip']
        return None

    def load_storage(self):
        model = LoadInfoModel()
        info = model.get_load_info()
        if info:
            data  = json.loads(info[0])
        else:
            data = []
        return data
